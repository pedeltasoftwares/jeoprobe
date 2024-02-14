import sys
import os
sys.path.append(os.getcwd())
from customtkinter import *
from PIL import Image, ImageTk
import tkinter
from tkinter import filedialog
import re
import numpy as np
import shutil
import xlsxwriter
import openpyxl

def generar_ventana_1():

    global ventana_nivel_1,textbox_ruta_carpeta_senales_peer,boton_siguiente_senales_peer,seleccionar_sismos

    #Ventana top level
    ventana_nivel_1 = CTkToplevel()
    #Nombre de la ventana
    ventana_nivel_1.title("Examinar")
    #Resizable
    ventana_nivel_1.resizable(False,False)
    ventana_nivel_1.transient(menu_window)
    ventana_nivel_1.grab_set()
    #Tema de la ventana
    set_appearance_mode("light")
    #Geometría
    width = 400
    height = 130
    screen_width = ventana_nivel_1.winfo_screenwidth()
    screen_height = ventana_nivel_1.winfo_screenheight()
    x = (screen_width - width) // 2
    y = (screen_height - height) // 2
    ventana_nivel_1.geometry(f"{width}x{height}+{x}+{y}")
    #Ícono ventana
    ventana_nivel_1.after(201, lambda :ventana_nivel_1.iconbitmap(os.path.join(images_path, "icono_principal.ico")))
    #Label
    label = CTkLabel(master=ventana_nivel_1,text=f"Ruta de almacenamiento señales PEER: ",font=('Gothic A1',13))
    label.place(x=20,y=5)
    #Textbox de la ruta de carpeta
    textbox_ruta_carpeta_senales_peer = CTkTextbox(master=ventana_nivel_1,font=('Gothic A1',12), text_color=("gray10", "gray90"),width=270,height=10,state=DISABLED,activate_scrollbars=False)
    textbox_ruta_carpeta_senales_peer.place(x=20,y=40)
    #Boton examinar
    boton_examinar_ruta = CTkButton(master= ventana_nivel_1, text="Examinar", width=80, height=12, compound="left",font=('Gothic A1',12),corner_radius=5, border_spacing=6,text_color=("gray10", "gray90"),fg_color=("gray85", "gray15"),hover_color=("gray75", "gray25"),border_color="black",border_width=0.75,command=seleccionar_ruta_senales_peer)
    boton_examinar_ruta.place(x=300,y=40)
    #Boton siguiente señales peer
    boton_siguiente_senales_peer = CTkButton(master= ventana_nivel_1, text="Siguiente", width=80, height=12, compound="left",font=('Gothic A1',12),corner_radius=5, border_spacing=6,text_color=("gray10", "gray90"),fg_color=("gray85", "gray15"),hover_color=("gray75", "gray25"),border_color="black",border_width=0.75,state=tkinter.DISABLED,command=verificar_sismos_peer)
    boton_siguiente_senales_peer.place(x=160,y=85)

def seleccionar_ruta_senales_peer():

    global textbox_ruta_carpeta_senales_peer, carpeta_senales,boton_siguiente_senales_peer

    carpeta_senales = filedialog.askdirectory()

    if carpeta_senales:

        # Muestra la ruta del archivo en el textbox
        textbox_ruta_carpeta_senales_peer.configure(state=NORMAL)
        textbox_ruta_carpeta_senales_peer.delete(0.0,"end")
        textbox_ruta_carpeta_senales_peer.insert(0.0,carpeta_senales)
        textbox_ruta_carpeta_senales_peer.configure(state=DISABLED)

        #Habilita el botón de siguiente
        boton_siguiente_senales_peer.configure(state=NORMAL)

def verificar_sismos_peer():

    def destruir_errores_ventana_1():
        #modifica el textbox
        textbox_ruta_carpeta_senales_peer.configure(state=NORMAL)
        textbox_ruta_carpeta_senales_peer.delete(0.0,"end")
        textbox_ruta_carpeta_senales_peer.configure(state=DISABLED)
        #destruye la ventana
        window_logs.destroy()


    #Warning si la carpeta está vacía
    if len(os.listdir(carpeta_senales)) == 0:

        window_logs = CTkToplevel()
        window_logs.title("Error")
        window_logs.resizable(False,False)
        window_logs.transient(ventana_nivel_1)
        window_logs.grab_set()
        width = 200
        height = 100
        screen_width = window_logs.winfo_screenwidth()
        screen_height = window_logs.winfo_screenheight()
        x = (screen_width - width) // 2
        y = (screen_height - height) // 2
        window_logs.geometry(f"{width}x{height}+{x}+{y}")
        #Ícono ventana
        window_logs.after(201, lambda :window_logs.iconbitmap(os.path.join(images_path, "error.ico")))
        #Label
        label_log = CTkLabel(master=window_logs,text=f"Carpeta vacía. \nPor favor seleccione otra.",font=('Gothic A1',13))
        label_log.place(x=30,y=18)
        #Botón de ok
        OKBoton_window_log = CTkButton(master= window_logs,text="OK", width=50, height=12, compound="left",font=('Gothic A1',12),corner_radius=5, border_spacing=6,text_color=("gray10", "gray90"),fg_color=("gray85", "gray15"),hover_color=("gray75", "gray25"),border_color="black",border_width=0.75, command=destruir_errores_ventana_1)
        OKBoton_window_log.place(x=80,y=65)
        
    else:
        contiene_AT2 = False
        #Warning si no hay extensiones correspondientes a .AT2
        for archivo in os.listdir(carpeta_senales):
            if archivo.endswith(".AT2"):
                contiene_AT2 = True
                break
        
        if contiene_AT2:
            # Minimizar la ventana actual
            ventana_nivel_1.withdraw()
            generar_ventana_2()

        else:
            window_logs = CTkToplevel()
            window_logs.title("Error")
            window_logs.resizable(False,False)
            window_logs.transient(ventana_nivel_1)
            window_logs.grab_set()
            width = 200
            height = 100
            screen_width = window_logs.winfo_screenwidth()
            screen_height = window_logs.winfo_screenheight()
            x = (screen_width - width) // 2
            y = (screen_height - height) // 2
            window_logs.geometry(f"{width}x{height}+{x}+{y}")
            #Ícono ventana
            window_logs.after(201, lambda :window_logs.iconbitmap(os.path.join(images_path, "error.ico")))
            #Label
            label_log = CTkLabel(master=window_logs,text=f"Archivos no válidos.",font=('Gothic A1',13))
            label_log.place(x=45,y=18)
            #Botón de ok
            OKBoton_window_log = CTkButton(master= window_logs,text="OK", width=50, height=12, compound="left",font=('Gothic A1',12),corner_radius=5, border_spacing=6,text_color=("gray10", "gray90"),fg_color=("gray85", "gray15"),hover_color=("gray75", "gray25"),border_color="black",border_width=0.75, command=destruir_errores_ventana_1)
            OKBoton_window_log.place(x=80,y=55)

def generar_ventana_2():

    global ventana_nivel_2,file_listbox

    #Abrir la ventana para seleccionar sismos
    ventana_nivel_2 = CTkToplevel()
    ventana_nivel_2.title("Seleccionar sismos")
    ventana_nivel_2.resizable(False,False)
    ventana_nivel_2.transient(ventana_nivel_1)
    ventana_nivel_2.grab_set()
    set_appearance_mode("light")
    width = 350
    height = 300
    screen_width = ventana_nivel_2.winfo_screenwidth()
    screen_height = ventana_nivel_2.winfo_screenheight()
    x = (screen_width - width) // 2
    y = (screen_height - height) // 2
    ventana_nivel_2.geometry(f"{width}x{height}+{x}+{y}")
    #Ícono ventana
    ventana_nivel_2.after(201, lambda :ventana_nivel_2.iconbitmap(os.path.join(images_path, "icono_principal.ico")))
    #Listbox
    file_listbox = tkinter.Listbox(ventana_nivel_2, selectmode=tkinter.MULTIPLE, width=50, height=15)
    file_listbox.place(x=22,y=10)
    files = list_files(carpeta_senales)
    file_listbox.delete(0, tkinter.END)
    for file in files:
        file_listbox.insert(tkinter.END, file)
    #Boton de volver
    boton_volver = CTkButton(master= ventana_nivel_2,text="Volver", width=50, height=12, compound="left",font=('Gothic A1',12),corner_radius=5, border_spacing=6,text_color=("gray10", "gray90"),fg_color=("gray85", "gray15"),hover_color=("gray75", "gray25"),border_color="black",border_width=0.75, command=f_volver) 
    boton_volver.place(x=80,y=265)
    #Boton de seleccionar
    boton_seleccionar = CTkButton(master= ventana_nivel_2,text="Escalar sismos", width=50, height=12, compound="left",font=('Gothic A1',12),corner_radius=5, border_spacing=6,text_color=("gray10", "gray90"),fg_color=("gray85", "gray15"),hover_color=("gray75", "gray25"),border_color="black",border_width=0.75, command=generar_ventana_3)
    boton_seleccionar.place(x=190,y=265)

def f_volver():

    #Destruye la ventana actual
    ventana_nivel_2.destroy()
    #Muestra la ventana anterior
    ventana_nivel_1.deiconify()

def list_files(directory):

    global earthquak_names

    #lee todos los archivos del directorio
    files = [file for file in os.listdir(directory) if file.endswith(".AT2")]
    #Extrae únicamente los nombres de los sismos
    earthquak_names = []
    for name in files:
        rgx = name.split("_")
        rgx = f"{rgx[0]}_{rgx[1]}"
        if rgx not in earthquak_names:
            earthquak_names.append(rgx)
        
    return earthquak_names

def generar_ventana_3():

    global input_factor_escala,escalar_sismos,boton_escalar,sismos_seleccionados,ventana_nivel_3

    sismos_seleccionados = [file_listbox.get(idx) for idx in file_listbox.curselection()]
    #TODO: borrar
    print(sismos_seleccionados)

    if len(sismos_seleccionados) == 0:
        window_logs = CTkToplevel()
        window_logs.title("Error")
        window_logs.resizable(False,False)
        window_logs.transient(ventana_nivel_2)
        window_logs.grab_set()
        width = 200
        height = 100
        screen_width = window_logs.winfo_screenwidth()
        screen_height = window_logs.winfo_screenheight()
        x = (screen_width - width) // 2
        y = (screen_height - height) // 2
        window_logs.geometry(f"{width}x{height}+{x}+{y}")
        #Ícono ventana
        window_logs.after(201, lambda :window_logs.iconbitmap(os.path.join(images_path, "error.ico")))
        #Label
        label_log = CTkLabel(master=window_logs,text=f"Debe seleccionar \nal menos un sismo.",font=('Gothic A1',13))
        label_log.place(x=45,y=18)
        #Botón de ok
        OKBoton_window_log = CTkButton(master= window_logs,text="OK", width=50, height=12, compound="left",font=('Gothic A1',12),corner_radius=5, border_spacing=6,text_color=("gray10", "gray90"),fg_color=("gray85", "gray15"),hover_color=("gray75", "gray25"),border_color="black",border_width=0.75, command=window_logs.destroy)
        OKBoton_window_log.place(x=80,y=60)

    else:

        #oculta la ventana de seleccion de sismos
        ventana_nivel_2.withdraw()

        #Abre la ventana para ingresar el factor para escalar sismos
        ventana_nivel_3 = CTkToplevel()
        ventana_nivel_3.title("Escalar sismos")
        ventana_nivel_3.resizable(False,False)
        ventana_nivel_3.transient(ventana_nivel_2)
        ventana_nivel_3.grab_set()
        set_appearance_mode("light")
        width = 200
        height = 120
        screen_width = ventana_nivel_3.winfo_screenwidth()
        screen_height = ventana_nivel_3.winfo_screenheight()
        x = (screen_width - width) // 2
        y = (screen_height - height) // 2
        ventana_nivel_3.geometry(f"{width}x{height}+{x}+{y}")
        #Ícono ventana
        ventana_nivel_3.after(201, lambda :ventana_nivel_3.iconbitmap(os.path.join(images_path, "icono_principal.ico")))
        #Label
        label = CTkLabel(master=ventana_nivel_3,text=f"Factor de escala:",font=('Gothic A1',13))
        label.place(x=50,y=8)
        #entry factor de escala
        factor_escala = ""
        input_factor_escala=CTkEntry(master=ventana_nivel_3, width=100, placeholder_text='Ej: 1.5')
        input_factor_escala.place(x=50, y=40)
        input_factor_escala.bind("<Leave>", lambda event: verificar_factor_escala())
        #Botón de escalar sismos
        boton_escalar = CTkButton(master= ventana_nivel_3,text="OK", width=50, height=12, compound="left",font=('Gothic A1',12),corner_radius=5, border_spacing=6,text_color=("gray10", "gray90"),fg_color=("gray85", "gray15"),hover_color=("gray75", "gray25"),border_color="black",border_width=0.75, state=tkinter.DISABLED, command=funcion_escalar_sismos)
        boton_escalar.place(x=75,y=80)

def verificar_factor_escala():
    global factor_escala

    factor_escala = input_factor_escala.get()

    patron = r'^[-+]?(\d+(\.\d*)?|\.\d+)$'
    match = re.match(patron, factor_escala) is not None and not factor_escala.isspace()

    if match:
        factor_escala = float(factor_escala)
        boton_escalar.configure(state=tkinter.NORMAL)
    else:

        window_logs = CTkToplevel()
        window_logs.title("Error")
        window_logs.resizable(False,False)
        window_logs.transient(ventana_nivel_3)
        window_logs.grab_set()
        width = 200
        height = 100
        screen_width = window_logs.winfo_screenwidth()
        screen_height = window_logs.winfo_screenheight()
        x = (screen_width - width) // 2
        y = (screen_height - height) // 2
        window_logs.geometry(f"{width}x{height}+{x}+{y}")
        #Ícono ventana
        window_logs.after(201, lambda :window_logs.iconbitmap(os.path.join(images_path, "error.ico")))
        #Label
        label_log = CTkLabel(master=window_logs,text=f"El factor de escala debe ser\nun valor número.",font=('Gothic A1',13))
        label_log.place(x=20,y=18)
        #Botón de ok
        OKBoton_window_log = CTkButton(master= window_logs,text="OK", width=50, height=12, compound="left",font=('Gothic A1',12),corner_radius=5, border_spacing=6,text_color=("gray10", "gray90"),fg_color=("gray85", "gray15"),hover_color=("gray75", "gray25"),border_color="black",border_width=0.75, command=window_logs.destroy)
        OKBoton_window_log.place(x=80,y=60)
        input_factor_escala.delete(0,"end")

def funcion_escalar_sismos():

    #Ruta de almacenamiento
    ruta_documentos = os.path.expanduser("~\\Documents")

    #Oculta la ventana del factor
    ventana_nivel_3.withdraw()

    if "log.log" in os.listdir(ruta_documentos):
        os.remove(f"{ruta_documentos}/log.log")

    def ejecutar_funcion_escalar_sismos():

        #inicializa el diccionario
        sismos = {}
        #Archivos en el directorio
        files_fullPath = os.listdir(carpeta_senales)
        #Lee los archivos seleccionados y los compila todos en un diccionario
        for file_name in sismos_seleccionados:
            sismos[file_name] = {} #inicializa el nombre del sisms
            for rooth_file in files_fullPath:
                if file_name in rooth_file and ".AT2" in rooth_file:
                    #lee el archivo
                    lines = read_earthquakefile(carpeta_senales,rooth_file)
                    #Formatea el archivo de sismo
                    vector, tiempo = format_earthquakefile(lines)
                    #Escala el sismo"
                    vector_escalado = [format(float(valor) * factor_escala, '.7E') for valor in vector]
                    #Almacena las aceleraciones
                    sismos[file_name][lines[1].split(",")[-1].split(" ")[1].split("\n")[0]] = {"vector":vector,"tiempo":tiempo,"vector_escalado":vector_escalado,"name_file": rooth_file.split(".AT2")[0]}
        
        #Escribe los archivos
        write_files(sismos)

        #Genera un excel para graficar los sismos seleccionados
        plot_signals(sismos,ruta_documentos)

        #Genera el log de terminación
        with open(f"{ruta_documentos}/log.log","w") as f:
            pass
        f.close()

    menu_window.after(100,ejecutar_funcion_escalar_sismos)
    generar_ventana_progreso_escalar_sismos()
        
def generar_ventana_progreso_escalar_sismos():

    window_logs_1 = CTkToplevel()
    window_logs_1.title("En progreso")
    window_logs_1.resizable(False,False)
    window_logs_1.transient(ventana_nivel_3)
    window_logs_1.grab_set()
    width = 250
    height = 80
    screen_width = window_logs_1.winfo_screenwidth()
    screen_height = window_logs_1.winfo_screenheight()
    x = (screen_width - width) // 2
    y = (screen_height - height) // 2
    window_logs_1.geometry(f"{width}x{height}+{x}+{y}")
    #Ícono ventana
    window_logs_1.after(1, lambda :window_logs_1.iconbitmap(os.path.join(images_path, "progreso.ico")))
    #Label
    label_log = CTkLabel(master=window_logs_1,text=f"Escalando sismos...",font=('Gothic A1',13))
    label_log.place(x=65,y=20)

    def destruir_ventana_progreso_escalar_sismos():

        # Ruta del archivo de log
        ruta_documentos = os.path.expanduser("~\\Documents")
        ruta_log = os.path.join(ruta_documentos, "log.log")

        if os.path.exists(ruta_log):

            window_logs_1.after(10, window_logs_1.destroy)

            window_logs = CTkToplevel()
            window_logs.title("Éxito")
            window_logs.resizable(False,False)
            window_logs.transient(ventana_nivel_3)
            window_logs.grab_set()
            width = 250
            height = 100
            screen_width = window_logs.winfo_screenwidth()
            screen_height = window_logs.winfo_screenheight()
            x = (screen_width - width) // 2
            y = (screen_height - height) // 2
            window_logs.geometry(f"{width}x{height}+{x}+{y}")
            #Ícono ventana
            window_logs.after(201, lambda :window_logs.iconbitmap(os.path.join(images_path, "exito.ico")))
            #Label
            label_log = CTkLabel(master=window_logs,text=f"Sismos escalados correctamente.",font=('Gothic A1',13))
            label_log.place(x=28,y=18)
            #Botón de ok
            OKBoton_window_log = CTkButton(master= window_logs,text="OK", width=50, height=12, compound="left",font=('Gothic A1',12),corner_radius=5, border_spacing=6,text_color=("gray10", "gray90"),fg_color=("gray85", "gray15"),hover_color=("gray75", "gray25"),border_color="black",border_width=0.75, command=window_logs.destroy)
            OKBoton_window_log.place(x=100,y=60)
        
        else:
            window_logs_1.after(1000, destruir_ventana_progreso_escalar_sismos)

    destruir_ventana_progreso_escalar_sismos()
    
def write_files(sismos:dict):

    #Ruta de almacenamiento
    ruta_documentos = os.path.expanduser("~\\Documents")

    #verifica si existe una carpeta llamada "input_deepsoil"
    for file in os.listdir(ruta_documentos):
        if "input_deepsoil" in file:
            shutil.rmtree(f"{ruta_documentos}\input_deepsoil")
    
    #Crea la carpeta que almacena el input de deep soil
    os.mkdir(f"{ruta_documentos}\input_deepsoil")
    os.mkdir(f"{ruta_documentos}\input_deepsoil/01_input")
    os.mkdir(f"{ruta_documentos}\input_deepsoil/02_consolidado/")
    os.mkdir(f"{ruta_documentos}\input_deepsoil/02_consolidado/original/")
    os.mkdir(f"{ruta_documentos}\input_deepsoil/02_consolidado/escalado/")

    #Escribe los archivos de entrada para deepsoil
    for sismo in list(sismos.keys()):
        #vector consolidado
        vector_consolidado = []
        for componente in list(sismos[sismo].keys()):
            nombre_archivo = sismos[sismo][componente]["name_file"] + ".txt"
            n_puntos = len(sismos[sismo][componente]["vector_escalado"])
            dt = sismos[sismo][componente]["tiempo"][1]
            cifras_decimales = contar_digitos_despues_del_punto(dt)
            with open(f"{ruta_documentos}\input_deepsoil/01_input/{nombre_archivo}","w") as f:
                #En la primera linea escribe el número de datos y el delta
                f.write(f"{n_puntos}\t{dt}\n")
                #Escribe la aceleración
                for i in range(len(sismos[sismo][componente]["vector_escalado"])):
                    t = round(float(sismos[sismo][componente]['tiempo'][i]),cifras_decimales)
                    a = sismos[sismo][componente]['vector_escalado'][i]
                    f.write(f"{t}\t{a}\n")
            f.close()
            vector_consolidado.append(sismos[sismo][componente]["vector_escalado"])
        
        #Escribe el vector consolidado escalado
        lines, _ = formatear_consolidado(sismos[sismo],"vector_escalado")

        with open(f"{ruta_documentos}\input_deepsoil/02_consolidado/escalado/{sismo}.txt", 'w') as f:
            f.write(f"t (s)\t{list(sismos[sismo].keys())[0]}\t{list(sismos[sismo].keys())[1]}\t{list(sismos[sismo].keys())[2]}\n")
            for line in lines:
                for value in line:
                    f.write(value)
                f.write("\n")

        f.close()

        #Escribe el vector consolidado original
        lines, _ = formatear_consolidado(sismos[sismo],"vector")

        with open(f"{ruta_documentos}\input_deepsoil/02_consolidado/original/{sismo}.txt", 'w') as f:
            f.write(f"t (s)\t{list(sismos[sismo].keys())[0]}\t{list(sismos[sismo].keys())[1]}\t{list(sismos[sismo].keys())[2]}\n")
            for line in lines:
                for value in line:
                    f.write(value)
                f.write("\n")

        f.close()
    
def plot_signals(sismos:dict,ruta_documentos:str):
    
    #Genera el excel
    wb = xlsxwriter.Workbook(f'{ruta_documentos}\input_deepsoil\output.xlsx')

    # Agregar una hoja de trabajo al archivo
    for sismo in list(sismos.keys()):
        # Establece el nombre de la hoja
        ws = wb.add_worksheet(sismo)
        #Genera el encabezado
        ws.merge_range('B2:E2', sismo, wb.add_format({'bold': True, 'align': 'center'}))
        #Establece ancho de columna
        ws.set_column('B:E', 15)
        #Centrar columnas
        formato_centrado = wb.add_format({'align': 'center'})
        #Genera el consolidado 
        lines,componentes = formatear_consolidado(sismos[sismo],"vector_escalado")
        #Escribe la primera linea
        ws.write('B3', 't(s)',formato_centrado), ws.write('C3', componentes[0],formato_centrado), ws.write('D3', componentes[1],formato_centrado), ws.write('E3', componentes[2],formato_centrado)
        #Escribe los datos
        row = 4
        col_index = ["B","C","D","E"]
        for line in lines:
            col = 0
            for value in line:
                #Extrae el valor antes del \t
                value = value.split('\t')[0]
                #Verifica que no sea espacio en blanco
                patron = re.compile(r'\S')
                if bool(patron.search(value)):
                    ws.write(f"{col_index[col]}{row}",float(value),formato_centrado)

                col+=1

            row+=1

        #Grafica
        chart = wb.add_chart({'type': 'scatter', 'subtype': 'straight'})
        chart.set_title({'name': sismo})
        chart.set_size({'width': 629, 'height': 399})  # Ancho y alto en píxeles
        chart.set_y_axis({'name': 'Aceleración (g)'})
        chart.set_x_axis({'name': 'tiempo (s)','position_axis': 'on_tick','major_gridlines': {'visible': True}, 'minor_gridlines': {'visible': True,'line': {'color': '#CCCCCC'}}})
        # Configurar los datos del gráfico de dispersión
        chart.add_series({
            'name': componentes[0],
            'categories': f'={sismo}!$B$4:$B${row-1}',  # Eje X
            'values':     f'={sismo}!$C$4:$C${row-1}',  # Eje Y
            'line': {'color': '#4472C4'}  # Color de la línea
        })
        chart.add_series({
            'name': componentes[1],
            'categories': f'={sismo}!$B$4:$B${row-1}',  # Eje X
            'values':     f'={sismo}!$D$4:$D${row-1}',  # Eje Y
            'line': {'color': '#ED7D31'}  # Color de la línea
        })
        chart.add_series({
            'name': componentes[2],
            'categories': f'={sismo}!$B$4:$B${row-1}',  # Eje X
            'values':     f'={sismo}!$E$4:$E${row-1}',  # Eje Y
            'line': {'color': '#7C7C7C'}
        })
        # Insertar el gráfico en la hoja de trabajo
        ws.insert_chart('H4', chart)
        
    wb.close()

def formatear_consolidado(sismo:dict,tipo_senal:str):

    #Obtiene la longitud del vector de tiempo para cada componente
    n_vector_tiempo = [len(sismo[componente]["tiempo"]) for componente in list(sismo.keys())]
    n_max = max(n_vector_tiempo)

    #nombres de las tres componentes
    componentes = list(sismo.keys())

    #delta de tiempo para redondear
    dt = sismo[componentes[0]]["tiempo"][1]
    cifras_decimales = contar_digitos_despues_del_punto(dt)

    #obtiene los tiempos y las aceleraciones de las tres componentes
    t1,a1 = sismo[componentes[0]]["tiempo"], sismo[componentes[0]][tipo_senal]
    t2,a2 = sismo[componentes[1]]["tiempo"], sismo[componentes[1]][tipo_senal]
    t3 ,a3 = sismo[componentes[2]]["tiempo"], sismo[componentes[2]][tipo_senal]

    #Genera el cuerpo del archivo
    lines = []
    for i in range(n_max):

        tmp_list = []

        #Agrega el tiempo
        try:
            tmp_list.append(f"{round(float(t1[i]),cifras_decimales)}\t")
        except IndexError:
            try:
               tmp_list.append(f"{round(float(t2[i]),cifras_decimales)}\t") 
            except IndexError:
                tmp_list.append(f"{round(float(t3[i]),cifras_decimales)}\t") 
        
        #Agrega las aceleraciones
        try:
            tmp_list.append(f"{a1[i]}\t")
        except IndexError:
            tmp_list.append(f" \t")
        
        try:
            tmp_list.append(f"{a2[i]}\t")
        except IndexError:
            tmp_list.append(f" \t")

        try:
            tmp_list.append(f"{a3[i]}\t")
        except IndexError:
            tmp_list.append(f" \t")

        lines.append(tmp_list)
        
    return lines,componentes

def read_earthquakefile(carpeta_senales:str,rooth_file:str):
    #Abre el archivo y lo lee
    with open(f"{carpeta_senales}/{rooth_file}","r") as f:
        lines = f.readlines()
    f.close()

    return lines

def format_earthquakefile(lines:str):

    file_info = lines[4:]

    #Lee las aceleraciones
    tmp_list,vector,tiempo = [], [], []
    for line in file_info:
        numeros = re.findall(r'[-+]?\d*\.?\d+E[+-]?\d+', line)
        tmp_list.append(numeros)
    
    for i in range(len(tmp_list)):
        for j in range(len(tmp_list[0])):
            try:
                vector.append(tmp_list[i][j])
            except IndexError:
                pass
    
    #Hace el vector de tiempo
    numeros = re.findall(r'[-+]?\d*\.\d+|\d+', lines[3])
    numeros = [int(numeros[0]),float(numeros[1])]
    vector_tiempo = np.arange(0, numeros[0] * numeros[1], numeros[1])
    tiempo = [str(valor) for valor in vector_tiempo]
    #elimina el último valor del tiempo siempre y cuando el tamaño del vector de aceleraciones no concuerde con el tamaño del vector de tiempo
    if len(vector) != len(tiempo):
        del tiempo[-1]
        
    return vector, tiempo

def contar_digitos_despues_del_punto(numero_str):
    # Buscar el índice del punto decimal en la cadena
    indice_punto = numero_str.find('.')
    
    # Si no se encuentra el punto decimal o está al final de la cadena, retornar 0
    if indice_punto == -1 or indice_punto == len(numero_str) - 1:
        return 0
    
    # Contar los dígitos después del punto
    digitos_despues_del_punto = len(numero_str) - indice_punto - 1
    return digitos_despues_del_punto

def generar_ventana_4():

    global ventana_nivel_4,textbox_ruta_output_deepsoil,boton_siguiente_output_deepsoil

    #Ventana top level
    ventana_nivel_4 = CTkToplevel()
    #Nombre de la ventana
    ventana_nivel_4.title("Examinar")
    #Resizable
    ventana_nivel_4.resizable(False,False)
    ventana_nivel_4.transient(menu_window)
    ventana_nivel_4.grab_set()
    #Tema de la ventana
    set_appearance_mode("light")
    #Geometría
    width = 400
    height = 130
    screen_width = ventana_nivel_4.winfo_screenwidth()
    screen_height = ventana_nivel_4.winfo_screenheight()
    x = (screen_width - width) // 2
    y = (screen_height - height) // 2
    ventana_nivel_4.geometry(f"{width}x{height}+{x}+{y}")
    #Ícono ventana
    ventana_nivel_4.after(201, lambda :ventana_nivel_4.iconbitmap(os.path.join(images_path, "icono_principal.ico")))
    #Label
    label = CTkLabel(master=ventana_nivel_4,text=f"Ruta de salida archivos DEEPSOIL: ",font=('Gothic A1',13))
    label.place(x=20,y=5)
    #Textbox de la ruta de carpeta
    textbox_ruta_output_deepsoil = CTkTextbox(master=ventana_nivel_4,font=('Gothic A1',12), text_color=("gray10", "gray90"),width=270,height=10,state=DISABLED,activate_scrollbars=False)
    textbox_ruta_output_deepsoil.place(x=20,y=40)
    #Boton examinar
    boton_examinar_ruta_output_deepsoil = CTkButton(master= ventana_nivel_4, text="Examinar", width=80, height=12, compound="left",font=('Gothic A1',12),corner_radius=5, border_spacing=6,text_color=("gray10", "gray90"),fg_color=("gray85", "gray15"),hover_color=("gray75", "gray25"),border_color="black",border_width=0.75,command=seleccionar_ruta_output_deepsoil)
    boton_examinar_ruta_output_deepsoil.place(x=300,y=40)
    #Boton siguiente señales peer
    boton_siguiente_output_deepsoil = CTkButton(master= ventana_nivel_4, text="Siguiente", width=80, height=12, compound="left",font=('Gothic A1',12),corner_radius=5, border_spacing=6,text_color=("gray10", "gray90"),fg_color=("gray85", "gray15"),hover_color=("gray75", "gray25"),border_color="black",border_width=0.75,state=tkinter.DISABLED,command=verificar_archivos_salidas_deepsoil)
    boton_siguiente_output_deepsoil.place(x=160,y=85)

def seleccionar_ruta_output_deepsoil():

    global carpeta_output_deepsoil

    carpeta_output_deepsoil = filedialog.askdirectory()

    if carpeta_output_deepsoil:

        # Muestra la ruta del archivo en el textbox
        textbox_ruta_output_deepsoil.configure(state=NORMAL)
        textbox_ruta_output_deepsoil.delete(0.0,"end")
        textbox_ruta_output_deepsoil.insert(0.0,carpeta_output_deepsoil)
        textbox_ruta_output_deepsoil.configure(state=DISABLED)

        #Habilita el botón de siguiente
        boton_siguiente_output_deepsoil.configure(state=NORMAL)

def verificar_archivos_salidas_deepsoil():

    def destruir_errores_ventana_4():
        #modifica el textbox
        textbox_ruta_output_deepsoil.configure(state=NORMAL)
        textbox_ruta_output_deepsoil.delete(0.0,"end")
        textbox_ruta_output_deepsoil.configure(state=DISABLED)
        #destruye la ventana
        window_logs.destroy()


    #Warning si la carpeta está vacía
    if len(os.listdir(carpeta_output_deepsoil)) == 0:

        window_logs = CTkToplevel()
        window_logs.title("Error")
        window_logs.resizable(False,False)
        window_logs.transient(ventana_nivel_4)
        window_logs.grab_set()
        width = 200
        height = 100
        screen_width = window_logs.winfo_screenwidth()
        screen_height = window_logs.winfo_screenheight()
        x = (screen_width - width) // 2
        y = (screen_height - height) // 2
        window_logs.geometry(f"{width}x{height}+{x}+{y}")
        #Ícono ventana
        window_logs.after(201, lambda :window_logs.iconbitmap(os.path.join(images_path, "error.ico")))
        #Label
        label_log = CTkLabel(master=window_logs,text=f"Carpeta vacía. \nPor favor seleccione otra.",font=('Gothic A1',13))
        label_log.place(x=30,y=18)
        #Botón de ok
        OKBoton_window_log = CTkButton(master= window_logs,text="OK", width=50, height=12, compound="left",font=('Gothic A1',12),corner_radius=5, border_spacing=6,text_color=("gray10", "gray90"),fg_color=("gray85", "gray15"),hover_color=("gray75", "gray25"),border_color="black",border_width=0.75, command=destruir_errores_ventana_4)
        OKBoton_window_log.place(x=80,y=65)
        
    else:
        contiene_xlsx = False
        #Warning si no hay extensiones correspondientes a .AT2
        for archivo in os.listdir(carpeta_output_deepsoil):
            if archivo.endswith(".xlsx"):
                contiene_xlsx = True
                break
        
        if contiene_xlsx:
            # Minimizar la ventana actual
            ventana_nivel_4.withdraw()
            compilar_output_deepsoil()

        else:
            window_logs = CTkToplevel()
            window_logs.title("Error")
            window_logs.resizable(False,False)
            window_logs.transient(ventana_nivel_4)
            window_logs.grab_set()
            width = 200
            height = 100
            screen_width = window_logs.winfo_screenwidth()
            screen_height = window_logs.winfo_screenheight()
            x = (screen_width - width) // 2
            y = (screen_height - height) // 2
            window_logs.geometry(f"{width}x{height}+{x}+{y}")
            #Ícono ventana
            window_logs.after(201, lambda :window_logs.iconbitmap(os.path.join(images_path, "error.ico")))
            #Label  
            label_log = CTkLabel(master=window_logs,text=f"Archivos no válidos.",font=('Gothic A1',13))
            label_log.place(x=45,y=18)
            #Botón de ok
            OKBoton_window_log = CTkButton(master= window_logs,text="OK", width=50, height=12, compound="left",font=('Gothic A1',12),corner_radius=5, border_spacing=6,text_color=("gray10", "gray90"),fg_color=("gray85", "gray15"),hover_color=("gray75", "gray25"),border_color="black",border_width=0.75, command=destruir_errores_ventana_4)
            OKBoton_window_log.place(x=80,y=55)

def compilar_output_deepsoil():

    #Ruta de almacenamiento
    ruta_documentos = os.path.expanduser("~\\Documents")

    #Oculta la ventana de la ruta
    ventana_nivel_4.withdraw()

    if "log_compilado.log" in os.listdir(ruta_documentos):
        os.remove(f"{ruta_documentos}/log_compilado.log")

    #verifica si existe una carpeta llamada "output_deepsoil"
    for file in os.listdir(ruta_documentos):
        if "output_deepsoil" in file:
            shutil.rmtree(f"{ruta_documentos}\output_deepsoil")

    try:
        os.mkdir(f"{ruta_documentos}\output_deepsoil")
    except OSError:
        pass

    def ejecutar_compilar_resultados_deepsoil():

        #Lee los archivos en la ruta
        file_list = os.listdir(carpeta_output_deepsoil)
        #A todos les quita el prefijo "Results_profile_0_motion_" y la extensión del archivo
        motion_name = [value.split("Results_profile_0_motion_")[1].split(".xlsx")[0] for value in file_list]

        #Inicializa diccionario de almacenamiento
        compilado_output_deepsoil = {}

        #Abre los archivos y extrae el PSA
        for f in motion_name:
            #Abre el libro
            wb = openpyxl.load_workbook(f"{carpeta_output_deepsoil}/Results_profile_0_motion_{f}.xlsx")
            #Selecciona las hojas
            layer1 = wb["Layer 1"]
            input_motion = wb["Input Motion"]
            # Crea el diccionario para almacenar
            compilado_output_deepsoil[f] = {"Layer1":[],"Input_motion":[]}
            #Encontra la máxima fila con datos en la columna del PSA en Layer1
            max_fila = encontrar_maxima_fila_convalor(layer1,"J")
            # Iterar sobre la hoja Layer 1 para extraer el PSA
            for fila in layer1.iter_rows(min_row=2, max_row=max_fila, min_col=10, max_col=10):
                for celda in fila:
                    compilado_output_deepsoil[f]["Layer1"].append(celda.value)
            #Encontra la máxima fila con datos en la columna del PSA en input motion
            max_fila = encontrar_maxima_fila_convalor(input_motion,"E")
            # Iterar sobre la hoja input motion para extraer el PSA
            for fila in input_motion.iter_rows(min_row=3, max_row=max_fila, min_col=5, max_col=5):
                for celda in fila:
                    compilado_output_deepsoil[f]["Input_motion"].append(celda.value)

            wb.close()

        #Generar un excel donde escribe el compilado
        wb = xlsxwriter.Workbook(f'{ruta_documentos}\output_deepsoil\output_deepsoil.xlsx')
        # Establece el nombre de la hoja a Layer 1
        ws = wb.add_worksheet("Layer 1")
        #Establece ancho de columna
        ws.set_column('A:ZZ', 25)
        #Centrar columnas
        formato_centrado = wb.add_format({'align': 'center'})
        #Escribe layer 1
        for j in range(len(list(compilado_output_deepsoil.keys()))):
            name_signal = list(compilado_output_deepsoil.keys())[j] #nombre del sismo
            ws.write(0,j,name_signal,formato_centrado)
            for i in range(len(compilado_output_deepsoil[name_signal]["Layer1"])):
                ws.write(i+1,j,float(compilado_output_deepsoil[name_signal]["Layer1"][i]),formato_centrado)

        # Establece el nombre de la hoja a Input Motion
        ws = wb.add_worksheet("Input Motion")
        #Establece ancho de columna
        ws.set_column('A:ZZ', 25)
        #Centrar columnas
        formato_centrado = wb.add_format({'align': 'center'})
        #Escribe layer 1
        for j in range(len(list(compilado_output_deepsoil.keys()))):
            name_signal = list(compilado_output_deepsoil.keys())[j] #nombre del sismo
            ws.write(0,j,name_signal,formato_centrado)
            for i in range(len(compilado_output_deepsoil[name_signal]["Input_motion"])):
                ws.write(i+1,j,float(compilado_output_deepsoil[name_signal]["Input_motion"][i]),formato_centrado)

        wb.close()
        
        #Genera el log de terminación
        with open(f"{ruta_documentos}/log_compilado.log","w") as f:
            pass
        f.close()

    menu_window.after(100,ejecutar_compilar_resultados_deepsoil)
    generar_ventana_progreso_compilar_resultados()

def encontrar_maxima_fila_convalor(hoja,columna):
    max_fila = hoja.max_row
    for fila in range(hoja.max_row, 1, -1):  # Iterar de la fila máxima hacia abajo
        if hoja[f"{columna}{fila}"].value is not None:  # Verificar si la celda en la columna J no está vacía
            max_fila = fila
            break

    return max_fila

def generar_ventana_progreso_compilar_resultados():

    window_logs_1 = CTkToplevel()
    window_logs_1.title("En progreso")
    window_logs_1.resizable(False,False)
    window_logs_1.transient(ventana_nivel_4)
    window_logs_1.grab_set()
    width = 250
    height = 80
    screen_width = window_logs_1.winfo_screenwidth()
    screen_height = window_logs_1.winfo_screenheight()
    x = (screen_width - width) // 2
    y = (screen_height - height) // 2
    window_logs_1.geometry(f"{width}x{height}+{x}+{y}")
    #Ícono ventana
    window_logs_1.after(1, lambda :window_logs_1.iconbitmap(os.path.join(images_path, "progreso.ico")))
    #Label
    label_log = CTkLabel(master=window_logs_1,text=f"Compilando resultados...",font=('Gothic A1',13))
    label_log.place(x=55,y=20)

    def destruir_ventana_progreso_compilar_resultados():

        # Ruta del archivo de log
        ruta_documentos = os.path.expanduser("~\\Documents")
        ruta_log = os.path.join(ruta_documentos, "log_compilado.log")

        if os.path.exists(ruta_log):

            window_logs_1.after(10, window_logs_1.destroy)

            window_logs = CTkToplevel()
            window_logs.title("Éxito")
            window_logs.resizable(False,False)
            window_logs.transient(ventana_nivel_4)
            window_logs.grab_set()
            width = 250
            height = 100
            screen_width = window_logs.winfo_screenwidth()
            screen_height = window_logs.winfo_screenheight()
            x = (screen_width - width) // 2
            y = (screen_height - height) // 2
            window_logs.geometry(f"{width}x{height}+{x}+{y}")
            #Ícono ventana
            window_logs.after(201, lambda :window_logs.iconbitmap(os.path.join(images_path, "exito.ico")))
            #Label
            label_log = CTkLabel(master=window_logs,text=f"Compilación de resultados exitosa.",font=('Gothic A1',13))
            label_log.place(x=28,y=18)
            #Botón de ok
            OKBoton_window_log = CTkButton(master= window_logs,text="OK", width=50, height=12, compound="left",font=('Gothic A1',12),corner_radius=5, border_spacing=6,text_color=("gray10", "gray90"),fg_color=("gray85", "gray15"),hover_color=("gray75", "gray25"),border_color="black",border_width=0.75, command=window_logs.destroy)
            OKBoton_window_log.place(x=95,y=60)
        
        else:
            window_logs_1.after(1000, destruir_ventana_progreso_compilar_resultados)

    destruir_ventana_progreso_compilar_resultados()

"""
MAIN

Nombres de las ventanas:
 - Ventana 1: Examinar la ruta de las señales.
 - Ventana 2: Se observan los nombres de los sismos. El usuario puede seleccionar los sismos.
 - Ventana 3: Incluir el factor de escala
 - Ventana 4: Seleccionar ruta para leer los archivos
"""
#Inicializa la ventana
menu_window = CTk()
#Geometría
width = 300
height = 200
screen_width = menu_window.winfo_screenwidth()
screen_height = menu_window.winfo_screenheight()
x = (screen_width - width) // 2
y = (screen_height - height) // 2
menu_window.geometry(f"{width}x{height}+{x}+{y}")
#Nombre de la ventana
menu_window.title("Procesamiento señales")
#Resizable
menu_window.resizable(False,False)
#Tema de la ventana
set_appearance_mode("light")
#Ícono ventana
images_path = os.path.join(os.path.dirname(os.path.realpath(__file__)), "images")
menu_window.after(201, lambda :menu_window.iconbitmap(os.path.join(images_path, "icono_principal.ico")))
#Botón principal para escalar señales
scale_signals_button = CTkButton(master= menu_window, corner_radius=5, height=40, border_spacing=10, text="Escalar señales",text_color=("gray10", "gray90"),fg_color=("gray85", "gray15"), image= CTkImage(Image.open(os.path.join(images_path, "icono_escalar.png")), size=(30, 30)), anchor="w",font=('Gothic A1',13),command=generar_ventana_1,hover_color=("gray75", "gray25"),border_color="black",border_width=0.75)
scale_signals_button.place(x=70, y=25)
#Botón principal para procesar resultados de deepsoil
postprocessing_deepsoil_button = CTkButton(master= menu_window, corner_radius=5, height=40, border_spacing=10, text="Compilar resultados DEEPSOIL",text_color=("gray10", "gray90"),fg_color=("gray85", "gray15"), image= CTkImage(Image.open(os.path.join(images_path, "icono_postprocesar.png")), size=(30, 30)), anchor="w",font=('Gothic A1',13),command=generar_ventana_4,hover_color=("gray75", "gray25"),border_color="black",border_width=0.75)
postprocessing_deepsoil_button.place(x=30, y=100)
#Label y logo de pedela
label = CTkLabel(master=menu_window,text=f"Desarrollado por: PEDELTA Colombia SAS ",font=('Gothic A1',9))
label.place(x=10,y=170)
#Ejecuta la ventana
menu_window.mainloop()
